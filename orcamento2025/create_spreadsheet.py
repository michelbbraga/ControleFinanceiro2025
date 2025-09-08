import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import sys
sys.path.append("/opt/.manus/.sandbox-runtime")
from data_api import ApiClient

def get_stock_price(symbol):
    client = ApiClient()
    try:
        response = client.call_api(
            "YahooFinance/get_stock_chart",
            query={
                "symbol": symbol,
                "region": "BR",
                "interval": "1d",
                "range": "1d",
                "includeAdjustedClose": True,
            },
        )
        if response and response["chart"]["result"]:
            # Get the latest adjusted close price
            latest_price = response["chart"]["result"][0]["indicators"]["adjclose"][0]["adjclose"][-1]
            return latest_price
    except Exception as e:
        print(f"Erro ao buscar preço para {symbol}: {e}")
    return None

def create_investment_spreadsheet(filename="controle_investimentos.xlsx"):
    workbook = openpyxl.Workbook()

    # Remover a aba padrão 'Sheet' criada pelo openpyxl
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Sheet: Ações
    acoes_sheet = workbook.create_sheet("Ações")
    acoes_sheet.append(["Nome do Ativo", "Nome da Empresa", "Data da Compra", "Quantidade", "Preço de Compra", "Preço Atual", "Rentabilidade Absoluta", "Rentabilidade Percentual"])
    acoes_sheet.append(["VALE3", "Vale S.A.", "2024-01-10", 50, 65.00, None, None, None]) # Preço Atual será atualizado
    acoes_sheet.append(["PETR4", "Petrobras", "2024-02-15", 100, 30.00, None, None, None]) # Preço Atual será atualizado

    # Atualizar preços das ações e calcular rentabilidade
    for row_num in range(2, acoes_sheet.max_row + 1):
        symbol = acoes_sheet[f'A{row_num}'].value
        if symbol:
            current_price = get_stock_price(symbol)
            if current_price:
                acoes_sheet[f'F{row_num}'].value = current_price # Atualiza Preço Atual
            else:
                acoes_sheet[f'F{row_num}'].value = acoes_sheet[f'E{row_num}'].value # Mantém Preço de Compra se não encontrar

        # Fórmulas de rentabilidade
        acoes_sheet[f'G{row_num}'] = f'=(F{row_num}-E{row_num})*D{row_num}'
        acoes_sheet[f'H{row_num}'] = f'=((F{row_num}-E{row_num})/E{row_num})*100'
        acoes_sheet[f'H{row_num}'].number_format = '0.00%'

    # Sheet: FII
    fii_sheet = workbook.create_sheet("FII")
    fii_sheet.append(["Nome do Ativo", "Nome da Empresa", "Data da Compra", "Quantidade", "Preço de Compra", "Preço Atual", "Rentabilidade Absoluta", "Rentabilidade Percentual"])
    fii_sheet.append(["MXRF11", "Maxi Renda FII", "2023-11-20", 200, 10.00, 10.50, None, None])
    fii_sheet.append(["HGLG11", "CSHG Logística", "2024-03-01", 50, 160.00, 158.00, None, None])
    for row_num in range(2, fii_sheet.max_row + 1):
        fii_sheet[f'G{row_num}'] = f'=(F{row_num}-E{row_num})*D{row_num}'
        fii_sheet[f'H{row_num}'] = f'=((F{row_num}-E{row_num})/E{row_num})*100'
        fii_sheet[f'H{row_num}'].number_format = '0.00%'

    # Sheet: Tesouro Direto
    tesouro_direto_sheet = workbook.create_sheet("Tesouro Direto")
    tesouro_direto_sheet.append(["Nome do Ativo", "Tipo do Ativo", "Data de Vencimento", "Tipo de Pagamento", "Data da Compra", "Quantidade", "Preço de Compra", "Preço Atual", "Rentabilidade Prevista", "Rentabilidade Absoluta", "Rentabilidade Percentual"])
    tesouro_direto_sheet.append(["Tesouro Selic 2027", "Pós-fixado", "2027-03-01", "Juros Semestrais", "2023-05-01", 1, 13000.00, 13500.00, "Selic + 0.05%", None, None])
    tesouro_direto_sheet.append(["Tesouro IPCA+ 2035", "Híbrido", "2035-05-15", "Principal", "2024-01-20", 0.5, 2500.00, 2600.00, "IPCA + 5.50%", None, None])
    for row_num in range(2, tesouro_direto_sheet.max_row + 1):
        tesouro_direto_sheet[f'J{row_num}'] = f'=(I{row_num}-H{row_num})*G{row_num}'
        tesouro_direto_sheet[f'K{row_num}'] = f'=((I{row_num}-H{row_num})/H{row_num})*100'
        tesouro_direto_sheet[f'K{row_num}'].number_format = '0.00%'

    # Sheet: Renda Fixa
    renda_fixa_sheet = workbook.create_sheet("Renda Fixa")
    renda_fixa_sheet.append(["Nome do Ativo", "Instituição", "Data de Vencimento", "Tipo de Pagamento", "Data da Compra", "Quantidade", "Preço de Compra", "Preço Atual", "Rentabilidade Prevista", "Rentabilidade Absoluta", "Rentabilidade Percentual"])
    renda_fixa_sheet.append(["CDB Banco X", "Banco X", "2025-10-20", "Principal", "2023-08-01", 1, 10000.00, 10800.00, "110% CDI", None, None])
    renda_fixa_sheet.append(["LCI Financeira Y", "Financeira Y", "2026-04-01", "Principal", "2024-02-01", 1, 5000.00, 5200.00, "98% CDI", None, None])
    for row_num in range(2, renda_fixa_sheet.max_row + 1):
        renda_fixa_sheet[f'J{row_num}'] = f'=(I{row_num}-H{row_num})*G{row_num}'
        renda_fixa_sheet[f'K{row_num}'] = f'=((I{row_num}-H{row_num})/H{row_num})*100'
        renda_fixa_sheet[f'K{row_num}'].number_format = '0.00%'

    # Sheet: Resumo Geral
    resumo_geral_sheet = workbook.create_sheet("Resumo Geral")
    resumo_geral_sheet.append(["Tipo de Ativo", "Rentabilidade Absoluta Total", "Rentabilidade Percentual Média"])

    # Coletar dados para o resumo geral
    data_for_summary = []

    # Ações
    total_rent_abs_acoes = 0.0
    total_rent_perc_acoes = 0.0
    count_acoes = 0
    for row_num in range(2, acoes_sheet.max_row + 1):
        preco_compra = acoes_sheet[f'E{row_num}'].value
        preco_atual = acoes_sheet[f'F{row_num}'].value
        quantidade = acoes_sheet[f'D{row_num}'].value
        if all(isinstance(v, (int, float)) for v in [preco_compra, preco_atual, quantidade]):
            rent_abs = (preco_atual - preco_compra) * quantidade
            rent_perc = ((preco_atual - preco_compra) / preco_compra) * 100
            total_rent_abs_acoes += rent_abs
            total_rent_perc_acoes += rent_perc
            count_acoes += 1
    data_for_summary.append(["Ações", total_rent_abs_acoes, total_rent_perc_acoes / count_acoes if count_acoes > 0 else 0])

    # FII
    total_rent_abs_fii = 0.0
    total_rent_perc_fii = 0.0
    count_fii = 0
    for row_num in range(2, fii_sheet.max_row + 1):
        preco_compra = fii_sheet[f'E{row_num}'].value
        preco_atual = fii_sheet[f'F{row_num}'].value
        quantidade = fii_sheet[f'D{row_num}'].value
        if all(isinstance(v, (int, float)) for v in [preco_compra, preco_atual, quantidade]):
            rent_abs = (preco_atual - preco_compra) * quantidade
            rent_perc = ((preco_atual - preco_compra) / preco_compra) * 100
            total_rent_abs_fii += rent_abs
            total_rent_perc_fii += rent_perc
            count_fii += 1
    data_for_summary.append(["FII", total_rent_abs_fii, total_rent_perc_fii / count_fii if count_fii > 0 else 0])

    # Tesouro Direto
    total_rent_abs_tesouro = 0.0
    total_rent_perc_tesouro = 0.0
    count_tesouro = 0
    for row_num in range(2, tesouro_direto_sheet.max_row + 1):
        preco_compra = tesouro_direto_sheet[f'H{row_num}'].value
        preco_atual = tesouro_direto_sheet[f'I{row_num}'].value
        quantidade = tesouro_direto_sheet[f'G{row_num}'].value
        if all(isinstance(v, (int, float)) for v in [preco_compra, preco_atual, quantidade]):
            rent_abs = (preco_atual - preco_compra) * quantidade
            rent_perc = ((preco_atual - preco_compra) / preco_compra) * 100
            total_rent_abs_tesouro += rent_abs
            total_rent_perc_tesouro += rent_perc
            count_tesouro += 1
    data_for_summary.append(["Tesouro Direto", total_rent_abs_tesouro, total_rent_perc_tesouro / count_tesouro if count_tesouro > 0 else 0])

    # Renda Fixa
    total_rent_abs_renda_fixa = 0.0
    total_rent_perc_renda_fixa = 0.0
    count_renda_fixa = 0
    for row_num in range(2, renda_fixa_sheet.max_row + 1):
        preco_compra = renda_fixa_sheet[f'H{row_num}'].value
        preco_atual = renda_fixa_sheet[f'I{row_num}'].value
        quantidade = renda_fixa_sheet[f'G{row_num}'].value
        if all(isinstance(v, (int, float)) for v in [preco_compra, preco_atual, quantidade]):
            rent_abs = (preco_atual - preco_compra) * quantidade
            rent_perc = ((preco_atual - preco_compra) / preco_compra) * 100
            total_rent_abs_renda_fixa += rent_abs
            total_rent_perc_renda_fixa += rent_perc
            count_renda_fixa += 1
    data_for_summary.append(["Renda Fixa", total_rent_abs_renda_fixa, total_rent_perc_renda_fixa / count_renda_fixa if count_renda_fixa > 0 else 0])

    for row_data in data_for_summary:
        resumo_geral_sheet.append(row_data)

    # Adicionar um gráfico de barras para comparar rentabilidade percentual geral
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Comparação de Rentabilidade Percentual por Tipo de Ativo"
    chart.y_axis.title = "Rentabilidade Percentual Média"
    chart.x_axis.title = "Tipo de Ativo"

    data = Reference(resumo_geral_sheet, min_col=3, min_row=2, max_row=resumo_geral_sheet.max_row, max_col=3)
    categories = Reference(resumo_geral_sheet, min_col=1, min_row=2, max_row=resumo_geral_sheet.max_row)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    resumo_geral_sheet.add_chart(chart, "E2")

    workbook.save(filename)
    print(f"Planilha \'{filename}\' criada com sucesso com cálculos e comparação de rentabilidade por tipo de ativo.")

if __name__ == "__main__":
    create_investment_spreadsheet()

